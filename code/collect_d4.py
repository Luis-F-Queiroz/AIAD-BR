#!/usr/bin/env python3
"""
D4 collection pipeline (v3) — full, automated candidate-surfacing for hand-coding.

Generalizes the v2 pilot (collect_d4_docs.py) to the whole codebook spec:
  18 core firms x FY2023 + FY2024 x the fixed document set.

It DOWNLOADS public filings and SURFACES AI/labor keyword passages with context;
it does NOT code anything (adopt_level, labor_signal etc. remain the human coder's
job). Sources, all public:
  - CVM IPE index (annual/integrated/sustainability report; Q4 earnings release)
      -> years 2024 + 2025 so both fiscal years are covered
  - CVM FRE dataset (Formulario de Referencia): document text  + structured
      employee headcount (codebook doc #1)
  - SEC EDGAR (20-F) for the one US-listed core firm, JBS N.V.

Text extraction: PyMuPDF, with Apple Vision OCR fallback (ocr_pdf.swift) for
image-based decks; XML for FRE; HTML for EDGAR. Resumable: cached .bin/.txt are
reused, so re-runs skip re-download and (slow) OCR.

Output:
  Data/d4_candidates.xlsx   sheets: candidates | coverage | headcount | firms
  Data/d4_candidates.csv    flat mirror of the candidates sheet
  Data/raw/d4_firms/<cd>/<fy>/  snapshots (+ .txt extraction cache)

Run with the interpreter that has PyMuPDF + openpyxl:
  /usr/local/bin/python3 Data/scripts/collect_d4.py
Optional args restrict to specific CVM codes: ... collect_d4.py 20575 19348
"""
import base64, csv, html, io, re, subprocess, sys, time, urllib.request, zipfile
from datetime import date
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl

ROOT = Path(__file__).resolve().parent.parent          # Data/
RAW = ROOT / "raw" / "d4_firms"
FRAME = ROOT / "d4_sampling_frame.csv"
OCR_SWIFT = ROOT / "scripts" / "ocr_pdf.swift"
UA = "polygence-research (luisqueiroz236@gmail.com)"
FISCAL_YEARS = [2023, 2024]
OCR_THRESHOLD = 500      # chars/page below which a PDF is treated as image-based
OCR_MAX_PAGES = 60
CTX = 220                # snippet context chars each side of a keyword hit
EDGAR_CIKS = {"20575": ["1791942", "1450123"]}  # JBS N.V. / JBS S.A.

# ---- keyword lists (codebook v0.2: PT + EN) --------------------------------
AI = [r"intelig[ê e]ncia artificial", r"\bIA\b", r"aprendizado de m[á a]quina",
      r"machine learning", r"aprendizado profundo", r"deep learning", r"IA generativa",
      r"generative ai", r"modelos? de linguagem", r"\bLLM\b", r"ci[ê e]ncia de dados",
      r"an[á a]lise preditiva", r"automa[ç c][ã a]o", r"\bRPA\b", r"chatbot",
      r"assistente (virtual|digital)", r"vis[ã a]o computacional",
      r"processamento de linguagem natural", r"\bPLN\b", r"artificial intelligence",
      r"data science", r"\bNLP\b", r"computer vision", r"algoritmo"]
LABOR = [  # PT
    r"efici[ê e]ncia", r"produtividade", r"colaboradores", r"funcion[á a]rios",
    r"headcount", r"efetivo", r"quadro de", r"reestrutura[ç c][ã a]o",
    r"redu[ç c][ã a]o de custos", r"otimiza[ç c][ã a]o",
    r"requalifica[ç c][ã a]o", r"capacita[ç c][ã a]o",
    r"contrata[ç c][ã a]o", r"vagas", r"demiss[ã a]o", r"postos de trabalho",
    # EN
    r"efficiency", r"productivity", r"workforce", r"staff", r"employees",
    r"restructuring", r"cost reduction", r"cost savings", r"cost-to-income",
    r"efficiency ratio", r"operating leverage", r"reskilling", r"upskilling",
    r"hiring", r"layoffs", r"job cuts", r"redundancies", r"workforce reduction",
    r"downsizing", r"redeployment", r"attrition"]
AI_RE = re.compile("|".join(AI), re.I)
LABOR_RE = re.compile("|".join(LABOR), re.I)

# IPE document slots (Categoria+Tipo+Assunto matched, in priority order).
# Matching runs on norm(): lowercased, hyphens/underscores -> spaces, so the real
# CVM label "Press-release" matches the "press release" pattern.
IPE_SLOTS = {
    "annual_report": [("relat", "integrado"), ("relato", "integrado"),
                      ("relat", "anual"), ("sustentabilidade",)],
    "results_release": [("apresenta", "resultado"), ("release de resultado",),
                        ("divulga", "resultado"), ("press release",),
                        ("gerencial",)],   # banks file earnings as "Relatorio de Analise Gerencial"
}
# Fiscal-year keying, per slot:
#   results_release -> Data_Referencia year == fy. The Q4/full-year release carries the
#     period end (fy-12-31); keying on filing year would also catch the Q1-Q3 releases.
#   annual_report -> Data_Referencia year == fy PRIMARY, with a guarded fallback to
#     Data_Entrega year == fy+1 (a report is always filed the year after the year it
#     covers). Most firms set Data_Referencia to the fiscal year-end and the primary
#     rule finds the real report; a minority date it by publication, and only those
#     need the fallback. The fallback excludes docs whose Data_Referencia is fy+1, so
#     it can never steal the *next* fiscal year's report.
# EXCLUDE drops announcement-type filings ("Lancamento do Relatorio Anual Integrado" is
# a notice that the report exists, not the report -- it matches the report patterns but
# carries almost no text).
EXCLUDE = re.compile(r"lan[çc]amento|aviso|convoca|errata|edital|agente fiduci", re.I)


def norm(s):
    return re.sub(r"\s+", " ", re.sub(r"[-_]", " ", (s or "").lower()))


# ---- fetch / extract -------------------------------------------------------
def fetch(url, tries=3, timeout=150):
    url = url.replace("http://", "https://")
    last = None
    for k in range(tries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read()
        except Exception as e:
            last = e
            time.sleep(3 * (k + 1))
    raise last


def ocr(path):
    try:
        p = subprocess.run(["swift", str(OCR_SWIFT), str(path), str(OCR_MAX_PAGES)],
                           capture_output=True, text=True, timeout=1800)
        return p.stdout or "[ocr-empty]"
    except Exception as e:
        return f"[ocr-error {e}]"


def garbled(text):
    """True if extracted text is real-length but gibberish (custom-font PDFs that
    map to wrong glyphs). Vowel-starved / few pronounceable words. Guards the OCR
    gate, which otherwise only fires on low CHAR COUNT and misses font garble."""
    t = text[:200000]
    letters = sum(c.isalpha() for c in t)
    if letters < 500:
        return False
    vowels = sum(c in "aeiouAEIOUáéíóúâêôãõà" for c in t)
    words = t.split()
    realish = sum(1 for w in words if 2 <= len(w) <= 18 and re.search(r"[aeiouáéíóú]", w, re.I))
    return vowels / letters < 0.28 or realish / max(len(words), 1) < 0.45


def pdf_text(path):
    try:
        doc = fitz.open(path)
        text = "\n".join(pg.get_text() for pg in doc)
        n = max(len(doc), 1)
        doc.close()
    except Exception as e:
        return f"[pdf-error {e}]", "error"
    if len(text) / n < OCR_THRESHOLD or garbled(text):
        return ocr(path), "vision-ocr"
    return text, "pymupdf"


def fre_text(blob):
    """The FRE WEB XML embeds the real document as base64-encoded PDF sections
    (65 MB of blobs for a big bank); the narrative is NOT present as XML text.
    Decode each blob and extract its text (PyMuPDF, text-only for speed)."""
    z = zipfile.ZipFile(io.BytesIO(blob))
    xml = z.read(next(n for n in z.namelist() if n.lower().endswith(".xml"))).decode("utf-8", "ignore")
    out = []
    for b in re.findall(r">([A-Za-z0-9+/]{500,}={0,2})<", xml):
        try:
            dec = base64.b64decode(b)
            if dec[:4] != b"%PDF":
                continue
            d = fitz.open(stream=dec, filetype="pdf")
            out.append("\n".join(pg.get_text() for pg in d))
            d.close()
        except Exception:
            continue
    return "\n".join(out)


def strip_markup(raw):
    """XML/HTML -> text: unescape entities (twice; FRE double-encodes) + drop tags."""
    try:
        s = raw.decode("utf-8", "ignore") if isinstance(raw, bytes) else raw
    except Exception:
        s = raw.decode("latin-1", "ignore")
    s = html.unescape(html.unescape(s))
    s = re.sub(r"(?is)<(script|style).*?</\1>", " ", s)
    s = re.sub(r"<[^>]+>", " ", s)
    return s


def to_text(path, blob):
    if blob[:4] == b"%PDF":
        return pdf_text(path)
    if blob[:2] == b"PK":                      # zip: pdfs (reports) or xml (FRE)
        try:
            z = zipfile.ZipFile(io.BytesIO(blob))
            names = z.namelist()
        except Exception as e:
            return f"[zip-error {e}]", "error"
        pdfs = [n for n in names if n.lower().endswith(".pdf")]
        if pdfs:
            out, how = [], "pymupdf"
            for n in pdfs:
                inner = path.parent / ("_" + Path(n).name)
                inner.write_bytes(z.read(n))
                t, h = pdf_text(inner)
                out.append(t)
                how = h if h != "pymupdf" else how
            return "\n".join(out), how
        xmls = [n for n in names if n.lower().endswith((".xml", ".html", ".htm"))]
        if xmls:
            t = fre_text(blob)                      # FRE = base64 PDF sections in XML
            if len(t) > 2000:
                return t, "fre-pdf"
            return "\n".join(strip_markup(z.read(n)) for n in xmls), "fre-xml"
        return "", "empty-zip"
    head = blob[:200].lstrip().lower()
    if head[:1] == b"<" or b"<html" in head:
        return strip_markup(blob), "html"
    try:
        return blob.decode("latin-1"), "raw"
    except Exception:
        return "", "error"


# control chars openpyxl refuses; they turn up in custom-font PDFs that extract garbled
ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037]")


def clean(v):
    return ILLEGAL.sub(" ", v) if isinstance(v, str) else v


def snippets(text):
    # drop base64/hex blobs (space-free runs >35 chars) so leaked binary can't
    # false-match the AI regex (e.g. "IA"/"llm" inside base64)
    text = re.sub(r"\S{36,}", " ", clean(text or ""))
    text = re.sub(r"\s+", " ", text)
    seen, out = set(), []
    for m in AI_RE.finditer(text):
        s, e = max(0, m.start() - CTX), min(len(text), m.end() + CTX)
        ctx = text[s:e].strip()
        key = ctx[:90]
        if key in seen:
            continue
        seen.add(key)
        out.append((m.group(0), bool(LABOR_RE.search(ctx)), "…" + ctx + "…"))
    return out


# ---- index loading (cached) ------------------------------------------------
def load_index_csv(name, url):
    """Ensure Data/raw/d4_firms/<name> exists (download+unzip if needed); return rows."""
    path = RAW / name
    if not path.exists():
        blob = fetch(url)
        if blob[:2] == b"PK":
            z = zipfile.ZipFile(io.BytesIO(blob))
            member = next(n for n in z.namelist() if n.lower() == name.lower())
            path.write_bytes(z.read(member))
        else:
            path.write_bytes(blob)
    return list(csv.DictReader(open(path, encoding="latin-1"), delimiter=";"))


def cvm_digits(s):
    return re.sub(r"\D", "", s or "").lstrip("0")


# ---- per-firm collection ---------------------------------------------------
def pick_ipe(rows_by_cd, cd, fy):
    rows = [r for r in rows_by_cd.get(cd, []) if not EXCLUDE.search(r["Assunto"] or "")]

    def match(subs, pats_list, datecol):
        subs = sorted(subs, key=lambda r: r[datecol] or "", reverse=True)
        for pats in pats_list:
            hit = next((r for r in subs
                        if all(p in norm(r["Categoria"] + " " + r["Tipo"] + " " + r["Assunto"])
                               for p in pats)), None)
            if hit:
                return hit
        return None

    picks = {}
    # results release: reference year == fiscal year
    rr = match([r for r in rows if (r["Data_Referencia"] or "")[:4] == str(fy)],
               IPE_SLOTS["results_release"], "Data_Referencia")
    if rr:
        picks["results_release"] = rr
    # annual report: reference year == fy, else filed in fy+1 (but not ref-dated fy+1,
    # so it can't steal the next fiscal year's report). CVM firms are inconsistent about
    # whether Data_Referencia is the fiscal year-end or the publication date, and several
    # agribusiness firms report on non-calendar (harvest) fiscal years, so a minority of
    # annual reports land in the adjacent fiscal-year slot or stay unmatched. The FRE
    # (36/36) and earnings release carry the AI disclosure regardless; the report is
    # supplementary. `doc_date`/`doc_filed` on every candidate let the coder verify.
    ar = match([r for r in rows if (r["Data_Referencia"] or "")[:4] == str(fy)],
               IPE_SLOTS["annual_report"], "Data_Referencia")
    if not ar:
        ar = match([r for r in rows if (r["Data_Entrega"] or "")[:4] == str(fy + 1)
                    and (r["Data_Referencia"] or "")[:4] != str(fy + 1)],
                   IPE_SLOTS["annual_report"], "Data_Entrega")
    if ar:
        picks["annual_report"] = ar
    return picks


def fre_for(fre_idx, fre_emp, cnpj, fy):
    """Return (link_doc, headcount_or_None) for the FRE covering fiscal year fy."""
    ds = fy + 1                                    # FRE dataset year (filed the next year)
    idx = [r for r in fre_idx.get(ds, []) if r["CNPJ_CIA"].strip() == cnpj]
    link = None
    if idx:
        latest = max(idx, key=lambda r: int(r["VERSAO"]))
        link = latest["LINK_DOC"]
    hc = None
    emp = [r for r in fre_emp.get(ds, []) if r["CNPJ_Companhia"].strip() == cnpj]
    if emp:
        vmax = max(int(r["Versao"]) for r in emp)
        rows = [r for r in emp if int(r["Versao"]) == vmax]
        try:
            hc = sum(int(float(r["Quantidade_Ate30Anos"] or 0))
                     + int(float(r["Quantidade_30a50Anos"] or 0))
                     + int(float(r["Quantidade_Acima50Anos"] or 0)) for r in rows)
        except Exception:
            hc = None
    return link, hc


def edgar_20f(cd, fy):
    """Best-effort: return (url, primary_doc_url) for JBS's 20-F covering fy, else None."""
    for cik in EDGAR_CIKS.get(cd, []):
        try:
            data = fetch(f"https://data.sec.gov/submissions/CIK{int(cik):010d}.json", timeout=40)
            import json
            rec = json.loads(data)["filings"]["recent"]
            for form, rpt, acc, doc in zip(rec["form"], rec["reportDate"],
                                           rec["accessionNumber"], rec["primaryDocument"]):
                if form == "20-F" and (rpt or "")[:4] == str(fy):
                    a = acc.replace("-", "")
                    return (f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={cik}",
                            f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{a}/{doc}")
        except Exception:
            continue
    return None


def get_doc_text(cd, fy, slot, url):
    """Download (cached) + extract (cached .txt). Returns (text, how, blob_present).

    The cache is keyed on the source URL (stored in a .url sidecar): if document
    selection later resolves this slot to a different filing, the stale .bin/.txt
    are discarded and refetched, rather than silently reused.
    """
    outdir = RAW / cd / str(fy)
    outdir.mkdir(parents=True, exist_ok=True)
    stem = outdir / slot
    binf, txtf, urlf = stem.with_suffix(".bin"), stem.with_suffix(".txt"), stem.with_suffix(".url")
    if urlf.exists() and urlf.read_text(encoding="utf-8").strip() != url.strip():
        for p in (binf, txtf):
            p.unlink(missing_ok=True)
    if txtf.exists():
        urlf.write_text(url, encoding="utf-8")   # adopt: label pre-sidecar caches
        return txtf.read_text(encoding="utf-8"), "cached", True
    blob = binf.read_bytes() if binf.exists() else None
    if blob is None:
        blob = fetch(url)
        binf.write_bytes(blob)
    text, how = to_text(binf, blob)
    txtf.write_text(text, encoding="utf-8")
    urlf.write_text(url, encoding="utf-8")
    return text, how, True


def main():
    only = set(sys.argv[1:])
    firms = [r for r in csv.DictReader(open(FRAME, encoding="utf-8"))
             if r["role"] == "core" and (not only or r["cd_cvm"] in only)]
    stamp = date.today().isoformat()

    ipe_rows = (load_index_csv("ipe_cia_aberta_2024.csv",
                "https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/IPE/DADOS/ipe_cia_aberta_2024.zip")
                + load_index_csv("ipe_cia_aberta_2025.csv",
                "https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/IPE/DADOS/ipe_cia_aberta_2025.zip"))
    rows_by_cd = {}
    for r in ipe_rows:
        rows_by_cd.setdefault(cvm_digits(r["Codigo_CVM"]), []).append(r)

    fre_idx, fre_emp = {}, {}
    for y in (2024, 2025):
        fre_idx[y] = load_index_csv(f"fre_cia_aberta_{y}.csv",
            f"https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/FRE/DADOS/fre_cia_aberta_{y}.zip")
        fre_emp[y] = load_index_csv(f"fre_cia_aberta_empregado_local_faixa_etaria_{y}.csv",
            f"https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/FRE/DADOS/fre_cia_aberta_{y}.zip")

    candidates, coverage, headcounts = [], [], []
    for f in firms:
        cd, cnpj, name = f["cd_cvm"], f["cnpj"].strip(), f["firm"]
        for fy in FISCAL_YEARS:
            found = {"annual_report": "", "results_release": "", "FRE": "", "form_20F": "n/a"}
            n_ai = n_lab = 0

            jobs = []  # (slot, doc_type, assunto, source_url, doc_date, doc_filed)
            for slot, r in pick_ipe(rows_by_cd, cd, fy).items():
                jobs.append((slot, f"{r['Categoria']} / {r['Tipo']}".strip(" /"),
                             r["Assunto"], r["Link_Download"],
                             r["Data_Referencia"], r["Data_Entrega"]))

            link, hc = fre_for(fre_idx, fre_emp, cnpj, fy)
            if hc is not None:
                headcounts.append({"firm": name, "cd_cvm": cd, "cnpj": cnpj,
                                   "fiscal_year": fy, "employees": hc, "source": "FRE RH table"})
            if link:
                jobs.append(("FRE", "Formulario de Referencia", "FRE WEB (strategy/risk)",
                             link, f"{fy}-12-31", ""))

            if cd in EDGAR_CIKS:
                found["form_20F"] = "missing"
                e = edgar_20f(cd, fy)
                if e:
                    jobs.append(("form_20F", "20-F", "SEC annual report (foreign issuer)",
                                 e[1], f"{fy}-12-31", ""))

            for slot, doc_type, assunto, url, doc_date, doc_filed in jobs:
                try:
                    text, how, _ = get_doc_text(cd, fy, slot, url)
                    sn = snippets(text)
                except Exception as ex:
                    print(f"  {name} FY{fy} {slot}: FAIL {ex}")
                    continue
                found[slot] = "yes" if slot not in found or found[slot] in ("", "missing", "n/a") else found[slot]
                if slot == "form_20F":
                    found["form_20F"] = "yes"
                for kw, lab, s in sn:
                    n_ai += 1
                    n_lab += int(lab)
                    candidates.append({
                        "firm": name, "cd_cvm": cd, "sector": f["sector"],
                        "size_tercile": f["size_tercile"], "fiscal_year": fy,
                        "doc_slot": slot, "doc_type": doc_type, "assunto": assunto,
                        "doc_date": doc_date, "doc_filed": doc_filed,
                        "source_url": url, "extraction": how, "keyword": kw,
                        "labor_cooccur": int(lab), "snippet": s, "access_date": stamp})
                print(f"  {name} FY{fy} {slot}: [{how}] {len(sn)} AI hits ({sum(l for _,l,_ in sn)} labor)")

            coverage.append({"firm": name, "cd_cvm": cd, "sector": f["sector"],
                             "fiscal_year": fy, "annual_report": found["annual_report"] or "missing",
                             "results_release": found["results_release"] or "missing",
                             "FRE": found["FRE"] or "missing", "form_20F": found["form_20F"],
                             "headcount": hc if hc is not None else "",
                             "n_ai_hits": n_ai, "n_labor_cooccur": n_lab})

    write_outputs(candidates, coverage, headcounts, firms)


def write_outputs(candidates, coverage, headcounts, firms):
    xlsx = ROOT / "d4_candidates.xlsx"
    wb = openpyxl.Workbook()
    def sheet(title, rows, cols):
        ws = wb.create_sheet(title)
        ws.append(cols)
        for r in rows:
            ws.append([clean(r.get(c, "")) for c in cols])
        ws.freeze_panes = "A2"
        for i, c in enumerate(cols, 1):
            w = max([len(str(c))] + [len(str(r.get(c, ""))) for r in rows[:200]] + [8])
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(w + 2, 80)
    wb.remove(wb.active)
    sheet("candidates", candidates,
          ["firm", "cd_cvm", "sector", "size_tercile", "fiscal_year", "doc_slot",
           "doc_type", "assunto", "doc_date", "doc_filed", "source_url", "extraction",
           "keyword", "labor_cooccur", "snippet", "access_date"])
    sheet("coverage", coverage,
          ["firm", "cd_cvm", "sector", "fiscal_year", "annual_report", "results_release",
           "FRE", "form_20F", "headcount", "n_ai_hits", "n_labor_cooccur"])
    sheet("headcount", headcounts,
          ["firm", "cd_cvm", "cnpj", "fiscal_year", "employees", "source"])
    sheet("firms", firms,
          ["sector", "size_tercile", "role", "firm", "legal_name", "cd_cvm", "cnpj"])
    wb.save(xlsx)

    csvp = ROOT / "d4_candidates.csv"
    if candidates:
        with csvp.open("w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(candidates[0].keys()))
            w.writeheader(); w.writerows(candidates)
    print(f"\nWrote {len(candidates)} candidates, {len(coverage)} firm-years, "
          f"{len(headcounts)} headcounts -> {xlsx.name} (+ .csv mirror)")


if __name__ == "__main__":
    main()
